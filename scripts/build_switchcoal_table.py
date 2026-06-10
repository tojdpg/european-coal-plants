#!/usr/bin/env python3
"""
Build the SwitchCoal conversion table (Europe + Türkiye) for the website.

SPINE  = SwitchCoal "Finance Model" workbook (SC_Europe_Sites_Copy) — carries the
         coal-to-clean conversion economics that GEM does NOT have. SwitchCoal's plant
         census is itself GEM Global Coal Plant Tracker, JANUARY 2023 release.
OVERLAY = GEM Global Coal Plant Tracker, JANUARY 2026 release (current status), joined
         per plant on coordinates (fallback: name) -> tells us which SwitchCoal sites
         are still running, already closed since 2023, or in reserve.

v1: SwitchCoal's own computed figures are carried 1:1 and labelled as such
    ("SwitchCoal-Originalwerte, GEM-2023-Basis"). No re-costing in v1.

Reads the two workbooks from the local Codex download snapshot (read-only).
Emits switchcoal-europe.json next to data.json. Does NOT touch data.json/index.html.
"""
import openpyxl, json, unicodedata, re, math, os, datetime
from collections import defaultdict

SNAP = "/Users/thorstenjelinek/Documents/Codex/2026-06-07/you-are-joining-the-wam-switchcoal/downloads/google-drive-switchcoal/files"
SC   = f"{SNAP}/SwitchCoal-Jan23-Data-2025-05-18-Battery.xlsx"
GEM  = f"{SNAP}/Global-Coal-Plant-Tracker-January-2026.xlsx"
OUT  = os.path.join(os.path.dirname(__file__), "..", "switchcoal-europe.json")

ISO2 = {
 "Bosnia and Herzegovina":"BA","Bosnia_and_Herzegovina":"BA","Bulgaria":"BG",
 "Czech Republic":"CZ","Czech_Republic":"CZ","Czechia":"CZ","Germany":"DE","Denmark":"DK",
 "Spain":"ES","Finland":"FI","France":"FR","Greece":"GR","Croatia":"HR","Hungary":"HU",
 "Italy":"IT","Moldova":"MD","Montenegro":"ME","North Macedonia":"MK","North_Macedonia":"MK",
 "Netherlands":"NL","Poland":"PL","Romania":"RO","Serbia":"RS","Slovenia":"SI","Slovakia":"SK",
 "Türkiye":"TR","Turkiye":"TR","Turkey":"TR","Ukraine":"UA","Kosovo":"XK","Slovak Republic":"SK",
 "United Kingdom":"GB","United_Kingdom":"GB","Ireland":"IE","Sweden":"SE","Belgium":"BE","Austria":"AT",
}
EXCLUDE = {"Russia","Russian Federation"}

def norm(s):
    if s is None: return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii","ignore").decode().lower()
    s = s.replace("power station","").replace("power plant","").replace("thermal","")
    return re.sub(r"[^a-z0-9]+"," ",s).strip()

def num(v):
    if isinstance(v,(int,float)): return float(v)
    if isinstance(v,str):
        try: return float(v.replace(",",""))
        except: return None
    return None

# ---------- GEM Jan-2026 -> plant index (worldwide, matched by coords) ----------
wb = openpyxl.load_workbook(GEM, read_only=True, data_only=True); ws = wb["Units"]
it = ws.iter_rows(values_only=True); H = {h:i for i,h in enumerate(next(it)) if h}
gem = defaultdict(lambda: {"name":None,"country":None,"sts":set(),"lat":None,"lon":None,"mw":0.0,"mw_op":0.0,"co2":0.0})
for row in it:
    p = row[H["Plant name"]];
    if not p: continue
    k = (row[H["Country/Area"]], norm(p)); d = gem[k]
    d["name"]=p; d["country"]=row[H["Country/Area"]]; d["sts"].add(row[H["Status"]])
    _cap = num(row[H["Capacity (MW)"]]) or 0.0
    d["mw"] += _cap
    if row[H["Status"]] == "operating": d["mw_op"] += _cap   # operating-only (excl. cancelled/retired phantom units)
    d["co2"] += num(row[H["Annual CO2 (million tonnes / annum)"]]) or 0.0
    la,lo = num(row[H["Latitude"]]), num(row[H["Longitude"]])
    if d["lat"] is None and la is not None: d["lat"],d["lon"]=la,lo
wb.close()
gem = [g for g in gem.values() if g["lat"] is not None]

LIVE={"operating"}; RESERVE={"mothballed","hibernating"}; BUILD={"construction","announced","permitted","pre-permit"}
def gem_match(lat,lon,name):
    if lat is None: return None
    best=None; bd=0.09
    nn=norm(name)
    for g in gem:
        d=math.hypot(g["lat"]-lat, g["lon"]-lon)
        if d<bd: bd=d; best=g
    if best is None:  # name fallback
        for g in gem:
            if norm(g["name"])==nn: return g
    return best

# ---------- SwitchCoal Europe spine ----------
C = dict(coaltype=1,country=3,subnat=4,plant=5,opex_avg=6,opex_man=7,opex_coal=8,
         loc=13,lat=14,lon=15,scstatus=17,trackerid=18,cap=24,egross=25,lifeco2=31,
         annco2=32,flh=39,pv_cost=75,pv_mw=76,wind_cost=88,wind_mw=91,stor_cost=101,
         bat_mw=109,wsb_inv=111,tprofit30=115,area=120,parent=131,owner=132,wiki=134,
         ren_price=146,profitable=147,proffactor=148)
wb = openpyxl.load_workbook(SC, read_only=True, data_only=True); ws = wb["SC_Europe_Sites_Copy"]
rows = list(ws.iter_rows(min_row=3, values_only=True)); wb.close()

def avail(sts):
    if sts & LIVE: return "running"        # still burning coal -> future conversion candidate
    if sts & RESERVE: return "reserve"     # mothballed/reserve
    if sts and sts <= {"retired","cancelled"}: return "closed"  # already shut -> available-now site
    return "unknown"

plants=[]; skipped=0
for r in rows:
    country = r[C["country"]]
    # whitelist real European countries (drops stray legend/parameter rows whose
    # country cell holds units like "kg / kWh", "1 = 100%", etc.)
    if not isinstance(country, str) or country.strip() not in ISO2 or country.strip() in EXCLUDE:
        skipped+=1; continue
    country = country.strip()
    lat,lon = num(r[C["lat"]]), num(r[C["lon"]])
    _nm = r[C["plant"]] or ""
    if "Mondi Steti" in _nm: lat,lon = 50.461, 14.376            # QA: wrong coords in spine
    elif _nm.startswith("Merkenich"): lat,lon = 50.99, 6.97
    g = gem_match(lat,lon, r[C["plant"]])
    if _nm.startswith("Merkenich"): g = None   # genuine GEM-2026 omission; corrected coords sit ~4km from a different (Chempark) plant
    gsts = sorted(g["sts"]) if g else []
    if g:
        gem_status = ("operating" if "operating" in g["sts"]
                      else "mothballed" if "mothballed" in g["sts"]
                      else gsts[0] if gsts else "absent")
    else:
        gem_status = "absent"
    # clean owner/parent/wiki: SwitchCoal stores a gem.wiki URL in 'Owner' and a GEM L-id in 'Parent'
    owner_raw, parent_raw, wiki_raw = r[C["owner"]], r[C["parent"]], r[C["wiki"]]
    wiki_val = wiki_raw.strip() if (isinstance(wiki_raw,str) and wiki_raw.strip().startswith("http")) else None
    owner_val = None
    if isinstance(owner_raw,str) and owner_raw.strip().startswith("http"):
        wiki_val = wiki_val or owner_raw.strip()                  # move misfiled wiki URL into wiki
    elif isinstance(owner_raw,str) and owner_raw.strip():
        owner_val = owner_raw.strip()
    parent_val = parent_raw if (isinstance(parent_raw,str) and not re.match(r'^L\d+$', parent_raw.strip())) else None

    # conversion economics, with data-quality guards
    cap = num(r[C["cap"]])
    repl_pv, repl_wind, repl_bat = num(r[C["pv_mw"]]), num(r[C["wind_mw"]]), num(r[C["bat_mw"]])
    invest, profit30, area = num(r[C["wsb_inv"]]), num(r[C["tprofit30"]]), num(r[C["area"]])
    ren_price = num(r[C["ren_price"]])
    dq = None
    repl_sum = (repl_pv or 0) + (repl_wind or 0)
    if cap and repl_sum and repl_sum / cap < 0.1:                 # corrupt source row (e.g. Kostolac): build ~100x too small
        repl_pv = repl_wind = repl_bat = invest = profit30 = area = None
        dq = "conversion figures unavailable (source-row error)"
    elif ren_price == 0:                                          # non-EU market: no offtake price -> profit is an artifact
        profit30 = None
        dq = "profit n/a — no market offtake price (non-EU)"

    rec = {
      "country": country, "iso2": ISO2.get(country,""),
      "subnational": r[C["subnat"]], "name": r[C["plant"]], "location": r[C["loc"]],
      "lat": lat, "lon": lon, "coal_type": r[C["coaltype"]], "gem_tracker_id": r[C["trackerid"]],
      # coal plant (SwitchCoal / GEM-2023 vintage)
      "capacity_mw": cap, "energy_gross_gwh": num(r[C["egross"]]),
      "annual_co2_mt": num(r[C["annco2"]]), "lifetime_co2_mt": num(r[C["lifeco2"]]),
      "full_load_hours": num(r[C["flh"]]), "opex_coal_ct_kwh": num(r[C["opex_coal"]]),
      # conversion build + economics (SwitchCoal originals)
      "repl_pv_mw": repl_pv, "repl_wind_mw": repl_wind, "repl_bat_mw": repl_bat,
      "pv_cost_ct_kwh": num(r[C["pv_cost"]]), "wind_cost_ct_kwh": num(r[C["wind_cost"]]),
      "storage_cost_ct_kwh": num(r[C["stor_cost"]]), "renewable_price_ct_kwh": ren_price,
      "wsb_invest_bn": invest, "total_profit_30y_bn": profit30,
      "profitability_eur_w": num(r[C["proffactor"]]), "switch_profitable": r[C["profitable"]],
      "area_need_km2": area, "data_quality": dq,
      "owner": owner_val, "parent": parent_val, "wiki": wiki_val,
      # SwitchCoal's own status (GEM-2023) + GEM-2026 overlay
      "sc_status": r[C["scstatus"]],
      "gem2026_status": gem_status,
      "gem2026_statuses": gsts,
      "gem2026_capacity_mw": (round(g["mw_op"],1) if g else None),   # operating-only
      "availability": avail(g["sts"]) if g else "untracked",
    }
    plants.append(rec)

# ---------- totals ----------
def s(key): return round(sum(p[key] for p in plants if p.get(key)),1)
by_country = defaultdict(int); by_avail = defaultdict(int); by_gem = defaultdict(int)
for p in plants:
    by_country[p["country"]]+=1; by_avail[p["availability"]]+=1; by_gem[p["gem2026_status"]]+=1
totals = {
  "n_plants": len(plants),
  "n_countries": len(by_country),
  "sum_capacity_mw": s("capacity_mw"),
  "sum_annual_co2_mt": s("annual_co2_mt"),
  "sum_repl_pv_mw": s("repl_pv_mw"), "sum_repl_wind_mw": s("repl_wind_mw"), "sum_repl_bat_mw": s("repl_bat_mw"),
  "sum_wsb_invest_bn": s("wsb_invest_bn"), "sum_total_profit_30y_bn": s("total_profit_30y_bn"),
  "by_country": dict(sorted(by_country.items())),
  "by_availability": dict(by_avail), "by_gem2026_status": dict(by_gem),
  "spine_source": "SwitchCoal Finance Model workbook (SC_Europe_Sites_Copy) = GEM Global Coal Plant Tracker January 2023 release + SwitchCoal conversion model (T. Schmidt / Kapica model, Nov 2023). Economics carried 1:1 (SwitchCoal-Originalwerte).",
  "status_overlay_source": "GEM Global Coal Plant Tracker, January 2026 release (Units sheet); joined per plant by coordinates (fallback name).",
  "scope_note": "SwitchCoal Europe sites, excl. Türkiye (not in SwitchCoal's Europe sheet), excl. Russia. SwitchCoal counts/CO2 are GEM-2023 vintage and include plants closed since; see gem2026_status / availability per row. gem2026_capacity_mw is operating-only. A few rows with corrupt source economics or non-EU markets carry a data_quality note.",
  "currency_note": "Investment & profit in SwitchCoal's original units (bn USD; profit over 30 years). Costs in ct/kWh as in the workbook.",
  "generated_from": [os.path.basename(SC), os.path.basename(GEM)],
  "generated_at": datetime.date.today().isoformat(),
}
out = {"plants": plants, "totals": totals}
with open(OUT,"w") as f: json.dump(out,f,ensure_ascii=False,indent=1)

# ---------- validation print ----------
print(f"WROTE {os.path.abspath(OUT)}")
print(f"plants: {len(plants)}  (skipped {skipped})  countries: {len(by_country)}")
print("by country:", dict(sorted(by_country.items(), key=lambda kv:-kv[1])))
print("availability overlay:", dict(by_avail))
print("gem2026 status overlay:", dict(by_gem))
print(f"sum capacity: {totals['sum_capacity_mw']:,} MW | sum annual CO2: {totals['sum_annual_co2_mt']:,} Mt")
print(f"sum repl PV {totals['sum_repl_pv_mw']:,} MW / Wind {totals['sum_repl_wind_mw']:,} MW / Bat {totals['sum_repl_bat_mw']:,} MW")
print(f"sum W-S-B invest {totals['sum_wsb_invest_bn']:,} bn | sum profit-30y {totals['sum_total_profit_30y_bn']:,} bn")
print("\nSAMPLE (3 DE + 2 PL):")
for p in plants:
    if p["name"] in ("Boxberg power station","Janschwalde power station","Buschhaus power station","Belchatow power station","Turów power station"):
        print(f"  {p['name'][:26]:26} {p['country'][:8]:8} cap={p['capacity_mw']} sc='{p['sc_status']}' gem2026='{p['gem2026_status']}' avail={p['availability']} "
              f"PV/W/B={p['repl_pv_mw']}/{p['repl_wind_mw']}/{p['repl_bat_mw']} inv={p['wsb_invest_bn']} prof30={p['total_profit_30y_bn']}")
